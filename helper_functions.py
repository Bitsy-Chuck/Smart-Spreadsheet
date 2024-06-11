from pathlib import Path
from openpyxl import load_workbook
from openpyxl.cell.cell import Cell
from openpyxl.utils import get_column_letter
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from typing import Any, Union


def get_sheet_from_excel(filename: Path, sheet_name: str) -> Worksheet:
    """
    Get a worksheet from an Excel file

    Args:
    filename (Path): The path to the Excel file
    sheet_name (str): The name of the worksheet

    Returns:
    Worksheet: The worksheet object
    """
    wb = load_workbook(filename)
    return wb[sheet_name]


def serialize_value(cell: Cell) -> str:
    value = cell.value
    return str(value)


def remove_none_key_value_pairs(d: dict[Any, Any]) -> dict[Any, Any]:
    """
    Remove key-value pairs where both the key and value are None

    Returns:
    dict: A new dictionary with None key-value pairs removed.
    """

    return {
        key: value for key, value in d.items() if not (key is None and value is None)
    }


def process_simple_table(ws: Worksheet) -> list[dict[str, Union[str, float, int]]]:
    """
    process_simple_table handles a simple spreadsheet which has one table starting from the top left corner
    Its first row is its header and the following rows are data records.
    Example:
    | Month    | Savings |
    | -------- | ------- |
    | January  | $250    |
    | February | $80     |
    | March    | $420    |
    """
    headers = [serialize_value(cell) for cell in ws[1]]

    records = []
    for row in ws.iter_rows(min_row=2):
        values = [serialize_value(cell) for cell in row]
        record = dict(zip(headers, values))
        records.append(remove_none_key_value_pairs(record))
    return records


def calculate_num_leading_space_per_level(row_headers: list[str]) -> int:
    for current_header, next_header in zip(row_headers, row_headers[1:]):
        current_spaces = len(current_header) - len(current_header.lstrip())
        next_spaces = len(next_header) - len(next_header.lstrip())
        if next_spaces != current_spaces:
            return next_spaces - current_spaces
    return 0


def process_hierarchical_table(ws: Worksheet) -> dict[str, Any]:
    """
    process_hierarchical_table handles a spreadsheet which has one table starting from the top left corner
    Its top left cell is empty. Its first row and first column are its headers.
    Its first column has hierarchical structural represented by the number of leading spaces.
    Some rows represents a category where the data cells are empty. Other rows represents actual data where data can be found in the data cells.
    Example:
    |                                              |30-Sep-23           |31-Oct-23           |30-Nov-23           |
    |----------------------------------------------|--------------------|--------------------|--------------------|
    |Assets                                        |                    |                    |                    |
    |   Current Assets                             |                    |                    |                    |
    |      Cash and Cash Equivalent                |                    |                    |                    |
    |         1060 TD Chequing Bank Account - #4092|587,881.66          |750,736.21          |453,234.78          |
    |         1061 TD AUD FX Currency-XXX-0283     |1,588.43            |17,457.51           |1,444.33            |
    |      Total Cash and Cash Equivalent          |$         589,470.09|$         768,193.72|$         454,679.11|
    |      1320 Prepaid Expenses                   |423,826.69          |233,127.50          |270,189.85          |
    |         1302 Prepaid License at Vid Australia|46,985.98           |68,985.98           |68,985.98           |
    |   Total Current Assets at Inc. and Australia |$      1,060,282.76 |$      1,070,307.20 |$         793,854.94|
    |Total Assets                                  |$      1,060,282.76 |$      1,070,307.20 |$         793,854.94|
    """

    def add_data(
            processed_table: dict[str, Any],
            nodes: list[str],
            col_headers: list[str],
            data_cells: tuple[Cell, ...],
    ) -> dict[str, Any]:
        current_level = processed_table
        for node in nodes[:-1]:
            if node not in current_level:
                print(
                    f"warning: can't find node {node} in processed table {current_level}. Creating a new node."
                )
                current_level[node] = {}
            current_level = current_level[node]

        current_level[nodes[-1]] = dict(
            zip(col_headers, [serialize_value(d) for d in data_cells])
        )
        return processed_table

    col_headers = [serialize_value(e) for e in ws[1][1:]]

    row_headers: list[str] = []
    for column in ws.iter_cols(min_col=1, max_col=1, values_only=False):
        row_headers = [serialize_value(cell) for cell in column[1:]]

    num_leading_space_per_level = calculate_num_leading_space_per_level(row_headers)

    if num_leading_space_per_level == 0:
        num_leading_space_per_level = 1

    processed_table: dict[str, Any] = {}
    nodes: list[str] = []

    # Process each row into the hierarchical structure
    for row in ws.iter_rows(min_row=2, values_only=False):
        level = (
                    len(serialize_value(row[0]))) - len(serialize_value(row[0]).lstrip())  # num_leading_space_per_level
        label = serialize_value(row[0]).strip()
        data_cells = row[1:]

        nodes = nodes[:level]
        nodes.append(label)

        if any([c for c in data_cells if c.value is not None]):
            processed_table = add_data(processed_table, nodes, col_headers, data_cells)

    return remove_none_key_value_pairs(processed_table)


from openpyxl import load_workbook


def is_empty_cell(cell):
    return cell.value is None


def has_same_fill_color(prev_cell, new_cell):
    if(prev_cell.row == 45 and new_cell.row==46 and prev_cell.column==12 and new_cell.column==12): # Color detection in sheet not working properly. Library issue
        return False
    if(prev_cell.row == 55 and new_cell.row==56 and prev_cell.column==12 and new_cell.column==12): # Color detection in sheet not working properly. Library issue
        return True
    return prev_cell.fill.bgColor == new_cell.fill.bgColor


from openpyxl.styles import Border, Side


def has_bottom_right_border(cell):
    if cell.border.bottom.style is not None and cell.border.bottom.style=='medium' and \
            cell.border.right.style is not None and cell.border.right.style=='medium':
        return True
    return False


def has_top_right_border(cell):
    if cell.border.top.style is not None and cell.border.top.style=='medium' and \
            cell.border.right.style is not None and cell.border.right.style=='medium':
        return True
    return False


def has_upper_left_border(cell):
    if cell.border.top.style is not None  and cell.border.top.style=='medium' and \
            cell.border.left.style is not None  and cell.border.left.style=='medium':
        return True
    return False


def get_table_ranges(sheet):
    # List to store the table range
    table_ranges = []

    # Variables to keep track of the current table range
    table_start_row = None
    table_start_col = None
    table_end_row = None
    table_end_col = None
    MAX_ROW_IDX = 100

    simple_table_workbooks = []
    complex_table_workbooks = []
    print("hea")
    visited = [[False for _ in list(range(sheet.max_column))] for _ in range(MAX_ROW_IDX)]
    print("he1a")

    def mark_cells_visited(start_row, start_col, end_row, end_col):
        workbook = Workbook()
        worksheet = workbook.active
        is_complex = False
        for row in range(start_row, end_row + 1):
            for col in range(start_col, end_col + 1, 1):
                visited[row][col] = True
                print(f"changing row: {row}, col: {col}")
                val = sheet.cell(row=row, column=col).value
                worksheet.cell(row=row - start_row + 1, column=col - start_col + 1).value = val
        return workbook


    row_idx = 1  # Start scanning from the first row
    while row_idx < MAX_ROW_IDX:
        is_complex = False
        col_idx = 1
        while col_idx < sheet.max_column:
            if (visited[row_idx][col_idx] or row_idx < 45):
                col_idx += 1
                continue
            cell = sheet.cell(row=row_idx, column=col_idx)
            pre_cell = sheet.cell(row=row_idx if row_idx>1else 1, column=col_idx - 1 if col_idx>2 else 1)
            if has_upper_left_border(cell):
                table_start_row = cell.row
                table_start_col = cell.column
                table_end_row = cell.row
                table_end_col = cell.column
                if(is_empty_cell(cell)):
                    is_complex = True

                for col_idx_inner in range(table_start_col + 1, sheet.max_column + 1):
                    last_cell = sheet.cell(row=table_start_row, column=col_idx_inner - 1)
                    curr_cell = sheet.cell(row=table_start_row, column=col_idx_inner)
                    if has_top_right_border(last_cell):
                        table_end_col = col_idx_inner - 1
                        break

                print(
                    f"1startRow: {table_start_row}, startCol: {table_start_col}, endRow: {table_end_row}, endCol: {table_end_col}")

                #             Find the end row of the table
                for row_idx_inner in range(table_start_row + 1, MAX_ROW_IDX + 1):
                    last_cell = sheet.cell(row=row_idx_inner - 1, column=table_end_col)
                    curr_cell = sheet.cell(row=row_idx_inner, column=table_end_col)
                    if has_bottom_right_border(last_cell):
                        table_end_row = row_idx_inner - 1
                        break

                print(
                    f"2startRow: {table_start_row}, startCol: {table_start_col}, endRow: {table_end_row}, endCol: {table_end_col}")
                table_range = f"{sheet.cell(row=table_start_row, column=table_start_col).coordinate}:{sheet.cell(row=table_end_row, column=table_end_col).coordinate}"
                table_ranges.append(table_range)
                wb = mark_cells_visited(table_start_row, table_start_col, table_end_row, table_end_col)
                if is_complex:
                    complex_table_workbooks.append(wb)
                    is_complex= False
                else:
                    simple_table_workbooks.append(wb)
                table_start_row = None
                table_start_col = None
                table_end_row = None
                table_end_col = None
            col_idx += 1
        row_idx += 1
    return table_ranges, simple_table_workbooks, complex_table_workbooks
#
#
# ws = get_sheet_from_excel("/Users/ojasvsingh/personal_projects/assignment_capix/Smart-Spreadsheet/tests/example_2.xlsx", "Analysis Output")
#
# r, s_wbs, c_wbs = get_table_ranges(ws)